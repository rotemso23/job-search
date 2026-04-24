"""
excel_helper.py — append jobs to job_tracker.xlsx
Usage:
    python excel_helper.py '[{"title": "...", "company": "...", "location": "...", "match": "strong|good|potential", "url": "..."}]'

match values: "strong", "good", "potential"
"""

import sys
import json
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

TRACKER_PATH = os.path.join(os.path.dirname(__file__), "job-results", "job_tracker.xlsx")
SHEET_NAME = "Job Search"

HEADER_FILL   = PatternFill(patternType="solid", fgColor="FF1F4E79")
HEADER_FONT   = Font(bold=True, size=11, color="FFFFFFFF")
HEADER_ALIGN  = Alignment(horizontal="center", vertical="center")

FILLS = {
    "strong":   PatternFill(patternType="solid", fgColor="FFC6EFCE"),
    "good":     PatternFill(patternType="solid", fgColor="FFFFEB9C"),
    "potential":PatternFill(patternType="solid", fgColor="FFFFCCCC"),
}

HEADERS = ["Role & Company", "Location", "Comments", "Applied?", "Link", "Heard Back?"]
COL_WIDTHS = [48, 22, 36, 14, 14, 16]

# Summary box sits in column H (col 8), one column gap after the main table
SUMMARY_COL = 8   # H
SUMMARY_LABEL_FILL  = PatternFill(patternType="solid", fgColor="FF1F4E79")
SUMMARY_LABEL_FONT  = Font(bold=True, size=11, color="FFFFFFFF")
SUMMARY_VALUE_FILL  = PatternFill(patternType="solid", fgColor="FFC6EFCE")
SUMMARY_VALUE_FONT  = Font(bold=True, size=14)

THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
DATA_ALIGN = Alignment(vertical="center", wrap_text=True)


def write_summary_box(ws):
    """Write/refresh the Applied count box in column H."""
    from openpyxl.utils import get_column_letter
    col = SUMMARY_COL
    col_letter = get_column_letter(col)

    # Row 1: label
    label = ws.cell(row=1, column=col, value="Applied")
    label.fill  = SUMMARY_LABEL_FILL
    label.font  = SUMMARY_LABEL_FONT
    label.alignment = Alignment(horizontal="center", vertical="center")
    label.border = BORDER
    ws.column_dimensions[col_letter].width = 14

    # Row 2: COUNTIF formula — counts "Yes" (case-insensitive) in the Applied? column (D)
    count = ws.cell(row=2, column=col, value='=COUNTIF(D2:D10000,"Yes")')
    count.fill  = SUMMARY_VALUE_FILL
    count.font  = SUMMARY_VALUE_FONT
    count.alignment = Alignment(horizontal="center", vertical="center")
    count.border = BORDER
    ws.row_dimensions[2].height = 32


def init_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.row_dimensions[1].height = 22
    for col, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
        ws.column_dimensions[cell.column_letter].width = width
    write_summary_box(ws)
    return wb, ws


def load_or_create():
    if os.path.exists(TRACKER_PATH):
        wb = load_workbook(TRACKER_PATH)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    else:
        wb, ws = init_workbook()
    return wb, ws


def append_jobs(jobs):
    wb, ws = load_or_create()
    for job in jobs:
        row = ws.max_row + 1
        fill = FILLS.get(job.get("match", "good").lower(), FILLS["good"])
        role_company = f"{job['title']} \u2014 {job['company']}"
        values = [role_company, job.get("location", ""), "", "", None, ""]
        ws.row_dimensions[row].height = 32
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = DATA_ALIGN
        # Col E: hyperlink
        link_cell = ws.cell(row=row, column=5, value="Open Job")
        link_cell.hyperlink = job.get("url", "")
        link_cell.fill = fill
        link_cell.border = BORDER
        link_cell.alignment = DATA_ALIGN
    write_summary_box(ws)
    # No freeze panes, scroll to top
    from openpyxl.worksheet.views import Selection
    ws.freeze_panes = None
    sv = ws.sheet_view
    sv.topLeftCell = "A1"
    sv.pane = None
    sv.selection = [Selection(activeCell="A1", sqref="A1")]
    wb.save(TRACKER_PATH)
    print(f"Appended {len(jobs)} job(s). Total rows: {ws.max_row - 1} (excl. header).")


def update_match(search_str, new_match):
    """Find rows where col A contains search_str and update their fill color."""
    wb, ws = load_or_create()
    fill = FILLS.get(new_match.lower(), FILLS["good"])
    updated = 0
    for row in ws.iter_rows(min_row=2, max_col=len(HEADERS)):
        cell_a = row[0]
        if cell_a.value and search_str.lower() in str(cell_a.value).lower():
            for cell in row:
                cell.fill = fill
            updated += 1
    if updated > 0:
        write_summary_box(ws)
        wb.save(TRACKER_PATH)
    print(f"Updated {updated} row(s) matching '{search_str}' to '{new_match}'.")
    return updated


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python excel_helper.py '<json array of jobs>'")
        sys.exit(1)
    if sys.argv[1] == "update":
        data = json.loads(sys.argv[2])
        update_match(data["search"], data["match"])
    else:
        jobs = json.loads(sys.argv[1])
        append_jobs(jobs)
